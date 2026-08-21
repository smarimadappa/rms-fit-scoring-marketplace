#!/usr/bin/env python3
"""
Build an RMS fit scorecard (.xlsx) from scoring JSON.

USAGE
    python scripts/build_scorecard.py <input.json> <output.xlsx>
    (run from the skill folder so ../weights.json resolves)

WHAT THIS DOES
The skill supplies each input's 0-100 sub_score. This script reads weights.json
and does ALL the math: renormalizes weights around unknown inputs, applies the
low-activity penalty, computes the final score, flags low confidence -- then
looks up the review-volume/pricing recommendation (the headline result; see
REVIEW VOLUME RECOMMENDATION below) and renders the scorecard. There is no
Go/Proceed with caution/No banding. Tuning weights.json or review_volume_bands.json
is therefore the only thing needed to retune the model.

INPUT JSON FORMAT
Do NOT pass weight, final_score, confidence_note, or review_volume_recommendation
-- all computed here. Input names MUST match weights_raw_points keys in
weights.json exactly; the script exits with an error naming any that don't.

Single product:
{
  "product": "Acme CRM",
  "product_id": "12345",
  "inputs": [
    {"input": "Market Presence Score", "raw": "62.4", "sub_score": 62.4,
     "source": "Looker 5042", "note": ""},
    {"input": "Account Segment", "raw": "Enterprise", "sub_score": 75,
     "source": "Looker 5043"},
    {"input": "Regional distribution", "raw": "unknown", "sub_score": null,
     "source": "web research", "note": "not found"}
    ...
  ],
  "recommendation": "free-text summary written by the skill"
}

Batch: {"products": [ <single-product object>, ... ]} -> Summary sheet plus one
detail sheet per product.

SPECIAL FLAGS
  "low_activity_top1": true   on the Customer industry input when the product's
                              #1 reviewer vertical is on the low-activity list.
                              Triggers the final-score penalty from weights.json.

  "service_disqualifier": true   on the PRODUCT object (top level, not an input)
                              when the resolved G2 listing is a service/provider,
                              not a software product (type != "Software"). Forces
                              final_score to 0 and the recommendation to a hard
                              disqualification message, regardless of any inputs
                              supplied -- pass "inputs": [] alongside it, since
                              Step 2 should never have run.

OPERATOR OVERRIDES (web-research inputs only)
Pass the operator's value as `raw` AND the researched value as `researched_raw`.
The scorecard then shows both, tags the source AUTHORITATIVE, and highlights the
row. Overridden inputs count as present and never count against confidence.
    {"input": "Total customers & end-users", "raw": "40,000 (contact-confirmed)",
     "researched_raw": "~12,000 (ZoomInfo est.)", "sub_score": 100,
     "source": "operator-supplied"}

REVIEW VOLUME RECOMMENDATION (the headline recommendation -- no Go/Proceed with
caution/No banding)
Pass "business_type" on the PRODUCT object: "new_business" or "existing_customer"
(from the skill's Salesforce Account.Type lookup keyed by G2 vendor_id -- see
SKILL.md). Omit or pass "unknown" if the lookup found no account/an unmapped Type;
the script then defaults to "new_business" and flags the note as unverified.

The script derives the segment bucket ("SMB & MM" vs "Commercial/Enterprise") from
the Account Segment input's sub_score -- no extra field needed. It then looks up
the score's band in review_volume_bands.json for that bucket + business_type and
attaches the result as product["review_volume_recommendation"] -- this IS the
recommendation shown in the scorecard, there is no separate Go/Proceed with
caution/No label. If Account Segment is unknown (or the product was
service-disqualified), the recommendation is marked unavailable/disqualified with
a reason instead of guessed. On low confidence (see weights.json), the real
package/price recommendation is still returned, just flagged "LOW CONFIDENCE" in
its notes -- it is never forced down to a lower tier.
    {"product": "Acme CRM", "business_type": "new_business", "inputs": [...]}
"""


import json
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load all tunables from the single source of truth (weights.json at skill root).
_WEIGHTS_PATH = os.path.join(os.path.dirname(__file__), "..", "weights.json")
try:
    with open(_WEIGHTS_PATH) as _f:
        _CFG = json.load(_f)
except (FileNotFoundError, json.JSONDecodeError) as _e:
    sys.exit(f"FATAL: cannot read weights.json at {_WEIGHTS_PATH} ({_e}). "
             "Run this script from the skill folder so it can find weights.json.")

_WEIGHTS_RAW = _CFG.get("weights_raw_points", {})
_CONF        = _CFG.get("low_confidence_rule", {})
_HIGH        = set(_CFG.get("high_weight_inputs", []))
_PENALTY     = _CFG.get("low_activity_top1_penalty", 0.7)

# Review-volume/packaging recommendation table (separate tunable file -- see its
# own header comments for the segment-bucket and business-type derivation rules).
_RV_PATH = os.path.join(os.path.dirname(__file__), "..", "review_volume_bands.json")
try:
    with open(_RV_PATH) as _f:
        _RV_CFG = json.load(_f)
except (FileNotFoundError, json.JSONDecodeError) as _e:
    sys.exit(f"FATAL: cannot read review_volume_bands.json at {_RV_PATH} ({_e}).")

_RV_BANDS = _RV_CFG.get("bands", {})


def segment_bucket_for(inputs):
    """Derive SMB & MM vs Commercial/Enterprise from the Account Segment sub_score."""
    seg = next((i for i in inputs if i.get("input") == "Account Segment"), None)
    if seg is None or seg.get("sub_score") is None:
        return None
    return "Commercial/Enterprise" if seg["sub_score"] >= 60 else "SMB & MM"


def review_volume_band_for(bucket, business_type, score):
    for band in _RV_BANDS.get(bucket, {}).get(business_type, []):
        if band["score_min"] <= score <= band["score_max"]:
            return band
    return None


def compute_review_volume_recommendation(product, low_conf=False):
    if product.get("service_disqualifier"):
        return {
            "packages": [{"label": "Not a fit for RMS — service/provider listing, not a software product", "price": "n/a"}],
            "notes": "",
            "disqualified": True,
            "unavailable_reason": None,
        }

    bucket = segment_bucket_for(product["inputs"])
    if bucket is None:
        return {"unavailable_reason": "Account Segment unknown -- cannot determine SMB/MM vs Commercial/Enterprise bucket"}

    business_type = product.get("business_type", "unknown")
    unverified = business_type not in ("new_business", "existing_customer")
    if unverified:
        business_type = "new_business"

    band = review_volume_band_for(bucket, business_type, product["final_score"])
    if band is None:
        return {"unavailable_reason": f"No review-volume band defined for {bucket} / {business_type} at score {product['final_score']}"}

    notes = band["notes"]
    if unverified:
        notes = (notes + " " if notes else "") + \
            "(business_type not supplied/resolved -- defaulted to new_business, verify against Salesforce.)"
    if low_conf:
        notes = (notes + " " if notes else "") + \
            "LOW CONFIDENCE -- verify inputs before pitching this package/price."

    return {
        "segment_bucket": bucket,
        "business_type": business_type,
        "packages": band["packages"],
        "notes": notes,
        "low_confidence": low_conf,
        "unavailable_reason": None,
    }


def compute(product):
    """Compute weights, final score, and confidence from sub-scores; look up the
    review-volume/pricing recommendation -- the headline result, no Go/Proceed
    with caution/No banding.

    The skill supplies each input's 0-100 sub_score (and `low_activity_top1`
    on the industry input when its #1 vertical is low-activity). All the math
    -- renormalizing weights around unknowns, the penalty, and the confidence
    flag -- happens here, reading weights.json, so tuning that one file changes
    every future scorecard with no hand-arithmetic.
    """
    # Hard disqualifier: a G2 service/provider listing, not a software product.
    # Bypasses every other input entirely -- this is not averaged or weighted.
    # Set by the skill at Step 1.5 when the resolved product's `type` isn't
    # "Software".
    if product.get("service_disqualifier"):
        for i in product.get("inputs", []):
            i["weight"] = 0
        product["final_score"] = 0
        product["confidence_note"] = "N/A — disqualified before scoring"
        product["review_volume_recommendation"] = compute_review_volume_recommendation(product)
        return product

    inputs = product["inputs"]

    # Fail loudly on a name that isn't in weights.json: a typo would otherwise
    # silently zero that input's weight and skew the score.
    unknown = [i["input"] for i in inputs if i["input"] not in _WEIGHTS_RAW]
    if unknown:
        sys.exit("FATAL: input name(s) not found in weights.json: "
                 + ", ".join(unknown)
                 + "\nNames must match weights_raw_points keys exactly.")

    known = [i for i in inputs if i.get("sub_score") is not None]
    total_raw = sum(_WEIGHTS_RAW.get(i["input"], 0) for i in known)
    for i in inputs:
        raw = _WEIGHTS_RAW.get(i["input"], 0)
        i["weight"] = 0 if (i.get("sub_score") is None or total_raw == 0) \
            else round(raw / total_raw * 100, 2)

    final = round(sum(i["sub_score"] * i["weight"] / 100 for i in known), 1)

    # Low-activity #1-vertical penalty: a dead top reviewer vertical is close to
    # disqualifying and must not be averaged away by otherwise-strong inputs.
    if any(i.get("low_activity_top1") for i in inputs):
        final = round(final * _PENALTY, 1)
        product["penalty_applied"] = f"x{_PENALTY} (low-activity #1 reviewer vertical)"
    product["final_score"] = final

    high_missing = sum(1 for i in inputs
                       if i["input"] in _HIGH and i.get("sub_score") is None)
    n_avail = len(known)
    low_conf = (high_missing >= _CONF.get("min_high_weight_missing", 2)
                or n_avail < _CONF.get("min_total_inputs_available", 7))

    product["confidence_note"] = "%d of %d inputs available%s" % (
        n_avail, len(inputs), " — LOW confidence" if low_conf else "")
    product["review_volume_recommendation"] = compute_review_volume_recommendation(product, low_conf)
    return product


FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=15)
LABEL_FONT = Font(name=FONT, bold=True, size=11)
BODY_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILLS = {
    "green": "C6EFCE",
    "amber": "FFEB9C",
    "red": "F4CCCC",
    "neutral": "F2F2F2",
}


def review_volume_fill(rv):
    """Color the recommendation by outcome: red = not a fit/disqualified,
    amber = low confidence, green = a real package/price, gray = unavailable."""
    rv = rv or {}
    if rv.get("unavailable_reason"):
        return PatternFill("solid", fgColor=STATUS_FILLS["neutral"])
    if rv.get("disqualified"):
        return PatternFill("solid", fgColor=STATUS_FILLS["red"])
    packages = rv.get("packages", [])
    if packages and packages[0].get("label", "").startswith("Not a strong fit"):
        return PatternFill("solid", fgColor=STATUS_FILLS["red"])
    if rv.get("low_confidence"):
        return PatternFill("solid", fgColor=STATUS_FILLS["amber"])
    return PatternFill("solid", fgColor=STATUS_FILLS["green"])


def write_detail(ws, p):
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"RMS Fit Scorecard — {p.get('product', 'Unknown')}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    pid = p.get("product_id", "")
    ws["A2"] = f"Product ID: {pid}" if pid else "Product ID: (not resolved)"
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="666666")
    ws.merge_cells("A2:F2")

    headers = ["Input", "Raw value", "Sub-score (0-100)", "Weight (%)",
               "Contribution", "Source / note"]
    hrow = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    r = hrow + 1
    for item in p.get("inputs", []):
        sub = item.get("sub_score")
        weight = item.get("weight", 0) or 0
        contrib = (sub * weight / 100.0) if sub is not None else None
        # Override handling: if researched_raw is present, this input was
        # operator-overridden. Show both values and tag the source authoritative.
        is_override = bool(item.get("researched_raw"))
        raw_display = item.get("raw", "")
        source_txt = item.get("source", "")
        if is_override:
            raw_display = f"{item.get('raw', '')}  [operator override; research found: {item.get('researched_raw')}]"
            source_txt = f"{source_txt or 'operator-supplied'} — AUTHORITATIVE (override)"
        note = " — ".join(x for x in [source_txt, item.get("note", "")] if x)
        row_vals = [
            item.get("input", ""),
            raw_display,
            "unknown" if sub is None else round(sub, 1),
            "—" if sub is None else round(weight, 2),
            "—" if contrib is None else round(contrib, 2),
            note,
        ]
        for c, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="left" if c in (1, 2, 6) else "center",
                vertical="center", wrap_text=(c in (2, 6)))
            if sub is None:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")
            elif is_override:
                cell.fill = PatternFill("solid", fgColor="E2EFDA")  # light green = operator-confirmed
        r += 1

    # Totals / result block
    r += 1
    ws.cell(row=r, column=1, value="FINAL SCORE").font = LABEL_FONT
    fs = ws.cell(row=r, column=3, value=round(p.get("final_score", 0), 1))
    fs.font = Font(name=FONT, bold=True, size=13)
    fs.alignment = Alignment(horizontal="center")

    r += 2
    ws.cell(row=r, column=1, value="RECOMMENDATION").font = LABEL_FONT
    r += 1
    rv = p.get("review_volume_recommendation") or {}
    rv_fill = review_volume_fill(rv)
    if rv.get("unavailable_reason"):
        cell = ws.cell(row=r, column=1, value=rv["unavailable_reason"])
        cell.font = Font(name=FONT, italic=True, size=9, color="666666")
        cell.fill = rv_fill
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    else:
        if rv.get("segment_bucket"):
            meta = ws.cell(row=r, column=1,
                            value=f"Segment: {rv.get('segment_bucket', '')}  |  Business type: {rv.get('business_type', '')}")
            meta.font = Font(name=FONT, italic=True, size=9, color="666666")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            r += 1
        for pkg in rv.get("packages", []):
            lbl = ws.cell(row=r, column=1, value=pkg.get("label", ""))
            lbl.font = Font(name=FONT, bold=True, size=11)
            lbl.fill = rv_fill
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            price = ws.cell(row=r, column=5, value=pkg.get("price", ""))
            price.font = Font(name=FONT, bold=True, size=11)
            price.fill = rv_fill
            price.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
            r += 1
        if rv.get("notes"):
            note = ws.cell(row=r, column=1, value=rv["notes"])
            note.font = Font(name=FONT, italic=True, size=9, color="666666")
            note.alignment = Alignment(wrap_text=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            r += 1

    r += 1
    ws.cell(row=r, column=1, value="Notes").font = LABEL_FONT
    r += 1
    note_cell = ws.cell(row=r, column=1, value=p.get("recommendation", ""))
    note_cell.font = BODY_FONT
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)

    r += 4
    conf = ws.cell(row=r, column=1, value=f"Confidence: {p.get('confidence_note', 'n/a')}")
    conf.font = Font(name=FONT, italic=True, size=9, color="666666")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    disc = ws.cell(row=r, column=1,
                   value="Score is directional. Weights and the review-volume/pricing table are provisional "
                         "and tunable (see references/scoring-model.md). Not a validated prediction of success.")
    disc.font = Font(name=FONT, italic=True, size=8, color="999999")
    disc.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=6)

    widths = [30, 22, 16, 12, 13, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def review_volume_summary_text(rv):
    rv = rv or {}
    if rv.get("unavailable_reason"):
        return rv["unavailable_reason"]
    text = " / ".join(f"{pkg.get('label', '')} ({pkg.get('price', '')})" for pkg in rv.get("packages", []))
    if rv.get("low_confidence"):
        text += "  [LOW CONFIDENCE]"
    return text


def write_summary(ws, products):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "RMS Fit Scoring — Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    headers = ["Product", "Final score", "Recommendation", "Confidence"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    r = 4
    for p in products:
        ws.cell(row=r, column=1, value=p.get("product", "")).font = BODY_FONT
        sc = ws.cell(row=r, column=2, value=round(p.get("final_score", 0), 1))
        sc.font = BODY_FONT
        sc.alignment = Alignment(horizontal="center")
        rv = p.get("review_volume_recommendation")
        rc = ws.cell(row=r, column=3, value=review_volume_summary_text(rv))
        rc.font = BODY_FONT
        rc.fill = review_volume_fill(rv)
        rc.alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=r, column=4, value=p.get("confidence_note", "")).font = BODY_FONT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    for i, w in enumerate([32, 12, 40, 24], start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def main():
    if len(sys.argv) != 3:
        print("Usage: python build_scorecard.py <input.json> <output.xlsx>")
        sys.exit(1)
    with open(sys.argv[1]) as f:
        data = json.load(f)

    wb = Workbook()
    products = data["products"] if "products" in data else [data]
    products = [compute(p) for p in products]

    if len(products) > 1:
        summary = wb.active
        summary.title = "Summary"
        write_summary(summary, products)
        for p in products:
            name = (p.get("product", "Product") or "Product")[:28]
            ws = wb.create_sheet(title=name)
            write_detail(ws, p)
    else:
        ws = wb.active
        ws.title = (products[0].get("product", "Scorecard") or "Scorecard")[:28]
        write_detail(ws, products[0])

    wb.save(sys.argv[2])
    print(f"Wrote {sys.argv[2]}")


if __name__ == "__main__":
    main()
